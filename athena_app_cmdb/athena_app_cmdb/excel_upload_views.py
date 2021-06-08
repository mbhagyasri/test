# Copyright Notice ===================================================
# This file contains proprietary information of CDK Global
# Copying or reproduction without prior written approval is prohibited.
# Copyright (c) 2021 =================================================


import pydash
import openpyxl
import logging
import json
import time


from django.conf import settings
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from django.views.decorators.cache import never_cache


from django.utils.decorators import method_decorator

from rest_framework.views import APIView


from rest_framework.response import Response
from . import models, serializers
from .operators import Validators
from .common import get_model

from libs.utility.commands import Execute
from libs.utility.config_parsers import get_config
from .utils.helper_methods import JsonUUIDEncoder
from .middleware import ViewException


logger = logging.getLogger(__name__)
config = get_config()
EXE = Execute()

IGNORE_FILTERS = ['page_size', 'page']
CACHE_TTL = getattr(settings, 'CACHE_TTL', DEFAULT_TIMEOUT)
FORMAT = 'json'


@method_decorator(never_cache, name='dispatch')
class LoadFromExcel(APIView):
    def post(self, request, ):

        errors = {}
        errors_list = []
        wb = {}
        if 'pod_template' not in request.FILES:
            ViewException(FORMAT, 'Excel file need to be submitted using key "pod_template".', 400)
        excel_file = request.FILES['pod_template']
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception as e:
            logger.exception(e)
            ViewException(FORMAT, str(e), 400)

        # process sheet via Import Control
        if 'ImportControl' not in wb:
            ViewException(FORMAT, "Expecting ImportControl sheet to process upload.  "
                                 "Please review the spreadsheet and try again", 406)
        # Processing Import Control
        import_control = self.process_import_control(request, wb['ImportControl'])
        logger.debug('processing Import Control')
        sheet_data = {}
        processed_sheets = []
        vcenters = []
        tasks_list = []
        response_message = None
        for item in import_control:
            logger.info('Processing sheet: {}'.format(item['sheet']))
            if item['sheet'] not in wb:
                ViewException(FORMAT, "Sheet {} not found for processing.".format(item['sheet']), 406)
            if item['sheet'] not in sheet_data:
                data = self.load_worksheet(wb[item['sheet']])
                sheet_data[item['sheet']] = data
            else:
                data = sheet_data[item['sheet']]
            objname = item['model']
            if not data:
                ViewException(FORMAT, "No data found for processing from sheet: {}".format(item['sheet']), 400)
            count = 0
            if item['data_set_num'] not in data:
                logger.error('No data found for section {}'.format(item['description']))
                return_data = {'Successfully Processed': processed_sheets,
                               'Current Failed Section': item['description'],
                               'Errors': "No data processed for this section. check spreadsheet data and try again"}
                return Response(return_data, status=400)
            for row in data[item['data_set_num']]:

                count = count + 1
                # Convert data into an object
                data_set = {}
                for key in row:
                    data_set = pydash.objects.set_(data_set, key, row[key])
                valid, rdata = self.process_data(request=request, objname=objname, data=data_set)
                if not valid:
                    # it failed
                    errors = {"name": row.get('name', ""), "path": "/{}".format(item['model']),
                              'Validation_Errors': rdata}
                    errors_list.append(errors)
                    if errors_list:
                        logger.error(errors_list)
                        return_data = {'Successfully Processed': processed_sheets,
                                       'Current Failed Section': item['description'], 'Errors': errors_list}
                        return Response(return_data, status=400)

                if count == 20:
                    time.sleep(1)
                    count = 0
            processed_sheets.append(item['description'])
            time.sleep(1)
            response_message = {"Message": "Spreadsheet was upload successfully.",
                                "Sections successfully processed": processed_sheets}
        if not response_message:
            response_message = {"Errors":
                                     "no data was imported.  Check that ImportControl sheet include "
                                     "TRUE values for items to be processed."}
            return Response(response_message, status=404)
        return Response(response_message, status=200)

    def process_data(self, request, objname, data):
        """
            Do a patch merge if exist or create a row in the database if the row does not exist that match the name
        :param request:
        :param objname: model name
        :param data: data object to process
        :return: [True|False, data] True if success and data.  False and validation error
        """
        logger.info('Processing {}: {}'.format(objname, data['name']))
        obj = get_model(objname)
        serializer_class = serializers.serializer_class_lookup[objname]
        # Does this object exists ?
        if obj.objects.filter(name=data['name']).exists():
            # retrieving this record
            obj = obj.objects.get(name=data['name'])
            obj_serializer = serializer_class(obj)
            # need to flatten to merge json.
            json_out = json.dumps(obj_serializer.data, cls=JsonUUIDEncoder)
            current_data = json.loads(json_out)
            # Perform PATCH / partial update.  We merge data with existing data
            rdata = EXE.data_merge(current_data, data)
            path = '/{}/:id'.format(objname)
            # Validate against update using put validation
            rdata = Validators().parse_mappings(data=rdata, path=path, method='put')
            if 'Validation_Errors' in rdata:
                return False, rdata
            serializer = serializer_class(obj, data=rdata)

        else:
            logger.info('{} with name {} does not exist.  Creating'.format(objname, data['name']))
            # We got here because this object doesn't exist yet
            path = '/{}'.format(objname)
            data = Validators().parse_mappings(data=data, path=path, method='post', raise_exception=False)
            if 'Validation Errors' in data:
                # Failed
                return False, data
            serializer = serializer_class(data=data)
        if serializer.is_valid():
            serializer.save()
            rdata = serializer.data
            return True, rdata
        else:
            return False, serializer.errors

    def process_import_control(self, request, ws):
        """
        Read In ImportControl and figure out which sheets to upload

        :param request: Request Data.  Need this object to process API responses on failure
        :type request: object
        :param ws: sheet data
        :param ws: worksheet data
        """
        return_data = []
        logger.info('Importing Import Control sheet')
        manual_select = self.check_manual_select(ws)
        sheet_data = self.load_worksheet(ws, data_for_athena_app_cmdb=False)
        # Now we need to process which sheets to process
        column_search = 'Select'
        if manual_select:
            column_search = 'ManualChoices'
        for row in sheet_data[1]:
            if column_search in row:
                data_def = {"sheet": row['Sheet'], "model": row['Destination'], "description": row['Description']}
                if row['Table']:
                    data_def["data_set_num"] = row['Table']
                else:
                    data_def["data_set_num"] = 0
                return_data.append(data_def)
        return return_data

    def check_manual_select(self, ws):
        manual_select = False
        max_col = ws.max_column
        max_row = ws.max_row
        for r in range(1, max_row + 1):
            for i in range(1, max_col + 1):
                # is this Manually Select cell
                cell = ws.cell(row=r, column=i)
                value = self.clean_string_value(cell.value)
                if value and isinstance(value, str) and value.lower() == 'manually select':
                    next_cell = ws.cell(row=r, column=(i + 1))
                    next_cell_value = self.clean_string_value(next_cell.value)
                    if next_cell_value and next_cell_value.lower() == "yes":
                        manual_select = True
                        return manual_select

        return manual_select

    def load_worksheet(self, ws, data_for_athena_app_cmdb=True):
        """
        Load the worksheet and separate in return data object of number of data set
        :param ws: wb worksheet
        :param data_for_athena_app_cmdb: Set to True if this sheet data is used for athena_app_cmdb load.  Will start reading from column name
                             forward.  If False, all columns are stored as data
        :return: return_data with each data set in the order read based on header row found starting from 0
        """
        keys = {}
        max_col = ws.max_column
        max_row = ws.max_row
        return_data = {}
        data_set_num = 1
        data_set = []
        need_header_row = 1
        ignore_data = True
        if data_for_athena_app_cmdb:
            name_found = 0
        else:
            name_found = 1
        local_keys = {}
        count = 0
        for r in range(1, max_row + 1):
            header_row = False
            row_data = {}
            need_reset = False
            count = count + 1
            for i in range(1, max_col + 1):
                cell = ws.cell(row=r, column=i)
                value = self.clean_string_value(cell.value)

                # check header_row.
                if need_header_row:
                    if i == 1 and value and str(value).lower() == "header":
                        # We found the header row
                        need_header_row = 0
                        header_row = True
                        name_found = 0 if data_for_athena_app_cmdb else 1
                        continue
                elif header_row:
                    # ignore column if name not found yet
                    if name_found == 1:
                        if not value:
                            continue
                        keys[i] = value
                    elif cell.value == "name":
                        keys[i] = cell.value
                        name_found = 1
                    elif name_found == 0:
                        continue
                    local_keys[i] = value

                elif not need_header_row and not ignore_data and i in local_keys and value:
                    # not header row and we need to load data into data set
                    row_data[local_keys[i]] = value
            # if this is the header row, we need to stuff it in for keys
            if header_row:
                # we finished reading header so the next row need to use as data row
                ignore_data = False
            elif row_data:
                # there is data to add to data_set
                data_set.append(row_data)
            elif not row_data and not need_header_row:
                need_reset = True
            # Are we at the last row ?
            if r == max_row:
                return_data[data_set_num] = data_set
            elif need_reset:
                # so this row is empty and we now need to store this dataset and increment the next dataset
                if data_set:
                    return_data[data_set_num] = data_set
                    data_set_num = data_set_num + 1
                    data_set = []
                need_header_row = 1
                ignore_data = True
                if data_for_athena_app_cmdb:
                    name_found = 0
                else:
                    name_found = 1
                local_keys = {}
            if count == 20:
                time.sleep(1)
                count = 0
        return return_data

    @staticmethod
    def clean_string_value(value):
        if value and value == "":
            return None
        elif isinstance(value, str) and value != "":
            value = value.strip()
            # Split comma delimited
            if "," in value and (not value.endswith('"') and not value.startswith('"')):
                value = value.split(',')
                array = []
                for v in value:
                    # only include it if there is a value for it
                    if v:
                        array.append(v.strip())
                value = array
        return value
